import openpyxl
import xlwings as xw

# Objective 1: Скопировать из source в conversion

# Загрузка исходного Excel-файла source.xlsx
source = openpyxl.load_workbook('../real_data/source.xlsx')
source_sheet_NA = source.worksheets[0]

# Загрузка файла conversion.xlsx
conversion = openpyxl.load_workbook('../real_data/conversion.xlsx')
conversion_sheet_paste = conversion.worksheets[0]

# Копирование диапазона E5:CK59 из source.xlsx NA в первый лист conversion.xlsx
source_range = source_sheet_NA['E5:CK59']
for i, row in enumerate(source_range):
    for j, cell in enumerate(row):
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            conversion_sheet_paste.cell(row=5 + i, column=5 + j, value=cell.value)

# Сохранение изменений в файле conversion.xlsx
conversion.save('../real_data/conversion.xlsx')

print("Objective 1 ✅")

# Objective 2: Использование xlwings для расчета и получения значений

# Открываем файл с помощью xlwings
wbxl = xw.Book('../real_data/conversion.xlsx')

# Выбираем второй лист
transposed = wbxl.sheets['Transposed']

# Извлечение значений из диапазона A4:Y122 после автоматического пересчета формул
calculated_values = transposed.range('A4:Y122').value

# Закрытие файла, сохранение не нужно, так как значения уже извлечены
wbxl.close()

# Загрузка файла database.xlsx
database = openpyxl.load_workbook('../real_data/database.xlsx')
database_sheet = database.active

# Поиск первой пустой строки в database.xlsx
def find_first_empty_row(sheet):
    for row in range(1, sheet.max_row + 1):
        if not any(cell.value for cell in sheet[row]):
            return row
    return sheet.max_row + 1

empty_row = find_first_empty_row(database_sheet)
print(empty_row)


# Вставка вычисленных значений в database.xlsx
for i, row in enumerate(calculated_values):
    for j, value in enumerate(row):
        database_sheet.cell(row=empty_row + i, column=3 + j, value=value)

# Сохранение изменений в файле database.xlsx
database.save('../real_data/database.xlsx')

print("Objective 2 ✅")